"""
Export logic: Excel (with colors, two sheets) and plain CSV.

Sheet 1 – "Liste de Débit"  : main cabinet parts
Sheet 2 – "Tiroirs"         : drawer box parts (Drw*)

Both sheets include a blank "Module" column (col B) that the user fills manually.
"""

import csv
from core.parser import decimal_to_fraction_str

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Piece-name routing ──────────────────────────────────────────────────────
# Piece names starting with these prefixes go to the drawer/tray sheet
DRW_PREFIXES = ("Drw", "Tray")


def _is_drw(name: str) -> bool:
    return any(name.startswith(p) for p in DRW_PREFIXES)


def split_piece_names(piece_names: list[str]) -> tuple[list[str], list[str]]:
    """Return (main_names, drawer_names) split from a flat piece_names list."""
    main = [p for p in piece_names if not _is_drw(p)]
    drw  = [p for p in piece_names if _is_drw(p)]
    return main, drw


# ── Formatting helpers ──────────────────────────────────────────────────────

def _hex_to_openpyxl_color(hex_color: str) -> str:
    color = hex_color.lstrip("#")
    if len(color) == 6:
        return "FF" + color.upper()
    return "FFFFFFFF"


def _fmt(value) -> str:
    if value == "" or value is None:
        return ""
    if isinstance(value, (int, float)):
        if isinstance(value, int):
            return str(value)
        return decimal_to_fraction_str(value)
    return str(value)


def _dimension_str(width, length) -> str:
    if width == "" or length == "" or width is None or length is None:
        return ""
    return f"{_fmt(width)} x {_fmt(length)}"


def _piece_entries(row_data: dict, piece_name: str) -> list:
    entries = row_data.get(f"{piece_name}_entries", [])
    if entries:
        return entries
    qty = row_data.get(f"{piece_name}_qty", "")
    if qty == "":
        return []
    return [{
        "qty": qty,
        "width": row_data.get(f"{piece_name}_w", ""),
        "length": row_data.get(f"{piece_name}_l", ""),
        "source_color": row_data.get(f"{piece_name}_color", "#FFFFFF"),
    }]


def _entries_fill(entries: list):
    colors = {
        e.get("source_color", "#FFFFFF")
        for e in entries
        if e.get("source_color")
    }
    colors.discard("#FFFFFF")
    if not colors:
        return None
    color_hex = next(iter(colors)) if len(colors) == 1 else "#FFE8CC"
    return PatternFill(
        start_color=_hex_to_openpyxl_color(color_hex),
        end_color=_hex_to_openpyxl_color(color_hex),
        fill_type="solid",
    )


# ── Sheet writer ────────────────────────────────────────────────────────────

def _write_sheet(
    ws,
    data: list,
    piece_names: list[str],
    sheet_title: str = "Liste de Débit",
    font_size: int = 14,
    col_scale: float = 0.75,
    qty_col_width: int = 5,
    dim_col_width: int = 18,
    cab_col_width: int = 10,
    mod_col_width: int = 14,
):
    """Write one pivot sheet into an openpyxl Worksheet.

    Column layout:
      A  = Meuble (cabinet ID)
      B  = Module (blank – user fills manually)
      C… = piece columns (qty + dimensions pairs)
    """
    header_font    = Font(bold=True, size=font_size)
    subheader_font = Font(bold=True, size=max(font_size - 1, 8))
    data_font      = Font(size=font_size)
    cab_font       = Font(bold=True, size=font_size)
    center_align   = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left_align     = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    header_fill    = PatternFill(start_color="FFD3D3D3", end_color="FFD3D3D3", fill_type="solid")
    subheader_fill = PatternFill(start_color="FFE8E8E8", end_color="FFE8E8E8", fill_type="solid")

    thin   = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Row 1: group headers ────────────────────────────────────────────
    # Col A: Meuble — smaller font so it doesn't widen the column
    meuble_font = Font(bold=True, size=max(font_size - 3, 9))
    c = ws.cell(row=1, column=1, value="Meuble")
    c.font = meuble_font; c.alignment = center_align
    c.fill = header_fill; c.border = border

    c = ws.cell(row=1, column=2, value="Module")
    c.font = header_font; c.alignment = center_align
    c.fill = header_fill; c.border = border

    col = 3
    for pname in piece_names:
        c = ws.cell(row=1, column=col, value=pname)
        c.font = header_font; c.alignment = center_align
        c.fill = header_fill; c.border = border
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
        # Apply font to the right (merged) cell too
        c2 = ws.cell(row=1, column=col + 1)
        c2.font = header_font; c2.fill = header_fill; c2.border = border
        col += 2

    # ── Row 2: sub-headers ──────────────────────────────────────────────
    for ci in [1, 2]:
        c = ws.cell(row=2, column=ci, value="")
        c.font = subheader_font; c.fill = subheader_fill; c.border = border

    col = 3
    for _ in piece_names:
        for sub in ["Qté", "Dimensions"]:
            c = ws.cell(row=2, column=col, value=sub)
            c.font = subheader_font; c.alignment = center_align
            c.fill = subheader_fill; c.border = border
            col += 1

    # ── Data rows — one sub-row per entry ──────────────────────────────
    ws_row = 3   # first data row (rows 1+2 are headers)
    for row_data in data:
        # Pre-collect entries for every piece
        all_entries = {p: _piece_entries(row_data, p) for p in piece_names}

        # Skip cabinets that have no entries for any column in this sheet
        if not any(all_entries.values()):
            continue

        max_rows = max((len(e) for e in all_entries.values()), default=1)
        max_rows = max(max_rows, 1)

        for sub_idx in range(max_rows):
            xl_row = ws_row + sub_idx

            if sub_idx == 0:
                c = ws.cell(row=xl_row, column=1, value=row_data.get("cabinet_id", ""))
                c.font = cab_font; c.alignment = left_align; c.border = border
                if max_rows > 1:
                    ws.merge_cells(start_row=xl_row, start_column=1,
                                   end_row=xl_row + max_rows - 1, end_column=1)
                    # Apply font to all merged cells in column A
                    for mr in range(xl_row + 1, xl_row + max_rows):
                        mc = ws.cell(row=mr, column=1)
                        mc.font = cab_font; mc.border = border

                c = ws.cell(row=xl_row, column=2, value="")
                c.font = data_font; c.border = border
                if max_rows > 1:
                    ws.merge_cells(start_row=xl_row, start_column=2,
                                   end_row=xl_row + max_rows - 1, end_column=2)
                    for mr in range(xl_row + 1, xl_row + max_rows):
                        mc = ws.cell(row=mr, column=2)
                        mc.font = data_font; mc.border = border

            col = 3
            for pname in piece_names:
                entries = all_entries[pname]

                if sub_idx < len(entries):
                    e    = entries[sub_idx]
                    fill = _entries_fill([e])
                    qty_text = _fmt(e.get("qty", ""))
                    dim_text = _dimension_str(e.get("width", ""), e.get("length", ""))
                else:
                    fill = None
                    qty_text = ""
                    dim_text = ""

                for val in [qty_text, dim_text]:
                    c = ws.cell(row=xl_row, column=col, value=val)
                    c.alignment = center_align; c.border = border
                    c.font = data_font
                    if fill:
                        c.fill = fill
                    col += 1

        ws_row += max_rows

    # ── Column widths — auto-fit based on content ──────────────────────
    ws.freeze_panes = "C3"

    # Cabinet column: auto-fit to longest ID
    max_cab_len = max(
        (len(str(row_data.get("cabinet_id", ""))) for row_data in data),
        default=4,
    )
    auto_cab_width = max(cab_col_width, max_cab_len * 1.2 + 1)
    ws.column_dimensions["A"].width = auto_cab_width
    ws.column_dimensions["B"].width = mod_col_width

    col = 3
    for pname in piece_names:
        ws.column_dimensions[get_column_letter(col)].width = qty_col_width

        # Dimension column: measure longest fraction string in this piece
        max_dim_len = len(pname)
        for row_data in data:
            for e in _piece_entries(row_data, pname):
                s = _dimension_str(e.get("width", ""), e.get("length", ""))
                if len(s) > max_dim_len:
                    max_dim_len = len(s)
        # Scale factor controls how tightly columns fit
        auto_dim = max(dim_col_width,
                       int(max_dim_len * (font_size / 11.0) * col_scale) + 1)
        ws.column_dimensions[get_column_letter(col + 1)].width = auto_dim
        col += 2

    ws.row_dimensions[1].height = font_size * 2.2
    ws.row_dimensions[2].height = font_size * 1.8
    # All data rows get a uniform height (one entry per physical row now)
    data_start = 3
    total_data_rows = sum(
        max(max((len(_piece_entries(rd, p)) for p in piece_names), default=1), 1)
        for rd in data
    )
    for r in range(data_start, data_start + total_data_rows):
        ws.row_dimensions[r].height = font_size * 1.6

    # ── Apply font_size to every cell that still has the default 11pt ──
    # Row 1 col 1 (Meuble header) intentionally uses a smaller font — skip it.
    meuble_coord = ws.cell(row=1, column=1).coordinate
    for row in ws.iter_rows():
        for cell in row:
            if cell.coordinate == meuble_coord:
                continue   # keep the intentional smaller font
            if cell.font is None or cell.font.size == 11.0:
                existing = cell.font
                cell.font = Font(
                    name=existing.name or "Calibri",
                    size=font_size,
                    bold=existing.bold,
                    italic=existing.italic,
                )


# ── Public export functions ─────────────────────────────────────────────────

def export_excel(data: list, piece_names: list, output_path: str,
                 font_size: int = 14, col_scale: float = 0.75,
                 qty_col_width: int = 5, dim_col_width: int = 18,
                 cab_col_width: int = 10, mod_col_width: int = 14):
    """Export to Excel with two sheets: main parts and drawer parts."""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl is required. Run: pip install openpyxl")

    main_names, drw_names = split_piece_names(piece_names)
    kw = dict(font_size=font_size, col_scale=col_scale,
              qty_col_width=qty_col_width, dim_col_width=dim_col_width,
              cab_col_width=cab_col_width, mod_col_width=mod_col_width)

    wb = openpyxl.Workbook()
    # Override the workbook default font size (Normal style = index 0)
    # so empty/merged cells inherit the chosen font size instead of 11pt.
    try:
        wb._named_styles[0].font = Font(name="Calibri", size=font_size)
    except Exception:
        pass  # non-critical — explicit cell fonts still apply

    ws1 = wb.active
    ws1.title = "Liste de Débit"
    _write_sheet(ws1, data, main_names, **kw)

    if drw_names:
        ws2 = wb.create_sheet(title="Tiroirs")
        _write_sheet(ws2, data, drw_names, sheet_title="Tiroirs", **kw)

    wb.save(output_path)


def export_csv(data: list, piece_names: list, output_path: str):
    """Export main (non-drawer) parts to CSV."""
    main_names, _ = split_piece_names(piece_names)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        row1 = ["Meuble", "Module"]
        for pname in main_names:
            row1.extend([pname, ""])
        writer.writerow(row1)

        row2 = ["", ""]
        for _ in main_names:
            row2.extend(["Qté", "Dimensions"])
        writer.writerow(row2)

        for row_data in data:
            row = [row_data.get("cabinet_id", ""), ""]
            for pname in main_names:
                entries = _piece_entries(row_data, pname)
                row.append("\n".join(_fmt(e.get("qty", "")) for e in entries))
                row.append("\n".join(
                    _dimension_str(e.get("width", ""), e.get("length", ""))
                    for e in entries
                ))
            writer.writerow(row)


def export_csv_drawers(data: list, piece_names: list, output_path: str):
    """Export drawer parts only to a separate CSV."""
    _, drw_names = split_piece_names(piece_names)
    if not drw_names:
        return

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)

        row1 = ["Meuble", "Module"]
        for pname in drw_names:
            row1.extend([pname, ""])
        writer.writerow(row1)

        row2 = ["", ""]
        for _ in drw_names:
            row2.extend(["Qté", "Dimensions"])
        writer.writerow(row2)

        for row_data in data:
            row = [row_data.get("cabinet_id", ""), ""]
            for pname in drw_names:
                entries = _piece_entries(row_data, pname)
                row.append("\n".join(_fmt(e.get("qty", "")) for e in entries))
                row.append("\n".join(
                    _dimension_str(e.get("width", ""), e.get("length", ""))
                    for e in entries
                ))
            writer.writerow(row)
